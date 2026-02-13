"""総務省選挙XLS/XLSXファイルパーサー.

XLSファイルの構造（小選挙区、1シート=1選挙区）:
    Row 1: 選挙日（例: 令和６年１０月２７日執行）
    Row 2: タイトル
    Row 3: 選挙区名（例: 北海道第１区）
    Row 4: 候補者名（各列に1名）
    Row 5: 政党名（各列に1名）
    Row 6+: 市区町村別得票数
    最終データ行: 合計得票数
"""

import logging
import re

from datetime import date
from pathlib import Path

from src.domain.value_objects.election_candidate import CandidateRecord, ElectionInfo
from src.infrastructure.importers._constants import PREFECTURE_NAMES


logger = logging.getLogger(__name__)

# 和暦→西暦変換
WAREKI_MAP = {
    "令和": 2018,
    "平成": 1988,
    "昭和": 1925,
}


def _zen_to_han(text: str) -> str:
    """全角数字を半角数字に変換する."""
    zen = "０１２３４５６７８９"
    han = "0123456789"
    table = str.maketrans(zen, han)
    return text.translate(table)


def _parse_wareki_date(text: str) -> date | None:
    """和暦の日付文字列を西暦dateに変換する.

    例: "令和６年１０月２７日執行" → date(2024, 10, 27)
    """
    if not text:
        return None
    text = _zen_to_han(str(text))
    pattern = r"(令和|平成|昭和)(\d+)年(\d+)月(\d+)日"
    match = re.search(pattern, text)
    if not match:
        return None
    era, year_str, month_str, day_str = match.groups()
    base_year = WAREKI_MAP.get(era)
    if base_year is None:
        return None
    year = base_year + int(year_str)
    return date(year, int(month_str), int(day_str))


def _extract_prefecture(district_name: str) -> str:
    """選挙区名から都道府県名を抽出する.

    例: "北海道第１区" → "北海道"
         "東京都第５区" → "東京都"
    """
    for pref in PREFECTURE_NAMES:
        if district_name.startswith(pref):
            return pref
    return ""


def _clean_cell_value(value: object) -> str | None:
    """セルの値を文字列にクリーンアップする."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    return s


def _parse_votes(value: object) -> int | None:
    """得票数セルの値をintに変換する."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = _zen_to_han(str(value).strip().replace(",", "").replace("，", ""))
    s = s.replace(".", "").replace(" ", "").replace("　", "")
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _sum_data_rows(rows: list[tuple[object, ...]], col_idx: int, start_row: int) -> int:
    """データ行の指定列を合算する（合計行がない場合のフォールバック）."""
    total = 0
    for row_idx in range(start_row, len(rows)):
        row = rows[row_idx]
        if col_idx >= len(row):
            continue
        votes = _parse_votes(row[col_idx])
        if votes is not None:
            total += votes
    return total


def parse_xls_file(
    file_path: Path,
) -> tuple[ElectionInfo | None, list[CandidateRecord]]:
    """XLS/XLSXファイルをパースして候補者データを抽出する.

    Args:
        file_path: XLS/XLSXファイルのパス

    Returns:
        (選挙情報, 候補者レコードのリスト)
    """
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx(file_path)
    elif suffix == ".xls":
        return _parse_xls(file_path)
    else:
        logger.error("未対応のファイル形式: %s", suffix)
        return None, []


def _parse_xlsx(file_path: Path) -> tuple[ElectionInfo | None, list[CandidateRecord]]:
    """openpyxlを使用して.xlsxファイルをパースする."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    all_candidates: list[CandidateRecord] = []
    election_info: ElectionInfo | None = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[tuple[object, ...]] = [
            tuple(cell for cell in row) for row in ws.iter_rows(values_only=True)
        ]
        if len(rows) < 5:
            logger.debug("シート '%s' のデータ行が不足: %d行", sheet_name, len(rows))
            continue

        candidates, sheet_election_info = _parse_rows(rows)
        if sheet_election_info and election_info is None:
            election_info = sheet_election_info
        all_candidates.extend(candidates)

    wb.close()
    return election_info, all_candidates


def _parse_xls(file_path: Path) -> tuple[ElectionInfo | None, list[CandidateRecord]]:
    """xlrdを使用して.xlsファイルをパースする."""
    import xlrd

    wb = xlrd.open_workbook(str(file_path))
    all_candidates: list[CandidateRecord] = []
    election_info: ElectionInfo | None = None

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        if ws.nrows < 5:
            logger.debug("シート '%s' のデータ行が不足: %d行", ws.name, ws.nrows)
            continue

        rows: list[tuple[object, ...]] = []
        for row_idx in range(ws.nrows):
            row: tuple[object, ...] = tuple(
                ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)
            )
            rows.append(row)

        candidates, sheet_election_info = _parse_rows(rows)
        if sheet_election_info and election_info is None:
            election_info = sheet_election_info
        all_candidates.extend(candidates)

    return election_info, all_candidates


def _find_total_row_index(rows: list[tuple[object, ...]], start_row: int) -> int | None:
    """合計行のインデックスを見つける."""
    for i in range(len(rows) - 1, start_row, -1):
        row = rows[i]
        if not row:
            continue
        first_cell = _clean_cell_value(row[0])
        if first_cell and first_cell in ("合計", "計", "合　計"):
            return i
        # 2番目のセルもチェック（レイアウトによる）
        if len(row) > 1:
            second_cell = _clean_cell_value(row[1])
            if second_cell and second_cell in ("合計", "計", "合　計"):
                return i
    return None


def _parse_rows(
    rows: list[tuple[object, ...]],
) -> tuple[list[CandidateRecord], ElectionInfo | None]:
    """行データから候補者レコードを抽出する.

    Args:
        rows: シートの全行データ

    Returns:
        (候補者リスト, 選挙情報)
    """
    if len(rows) < 5:
        return [], None

    # Row 1: 選挙日
    election_date_str = _clean_cell_value(rows[0][0]) if rows[0] else None
    election_date = _parse_wareki_date(election_date_str or "")

    # Row 3: 選挙区名（0-indexed: rows[2]）
    district_name = ""
    for cell in rows[2]:
        val = _clean_cell_value(cell)
        if val:
            district_name = _zen_to_han(val)
            break

    if not district_name:
        logger.debug("選挙区名が取得できません")
        return [], None

    prefecture = _extract_prefecture(district_name)

    # Row 4-5: 候補者名と政党名（0-indexed: rows[3], rows[4]）
    name_row = rows[3]
    party_row = rows[4]

    # 候補者データの列を特定（名前が入っている列）
    candidate_columns: list[int] = []
    for col_idx in range(len(name_row)):
        name = _clean_cell_value(name_row[col_idx])
        if name and not _is_header_cell(name):
            candidate_columns.append(col_idx)

    if not candidate_columns:
        logger.debug("候補者が見つかりません: %s", district_name)
        return [], None

    # 合計行を見つける
    total_row_idx = _find_total_row_index(rows, 5)

    # 候補者データを抽出
    candidates: list[CandidateRecord] = []
    for col_idx in candidate_columns:
        name = _clean_cell_value(name_row[col_idx])
        if not name:
            continue

        party_name = ""
        if col_idx < len(party_row):
            party_name = _clean_cell_value(party_row[col_idx]) or ""

        # 合計得票数
        total_votes = 0
        if total_row_idx is not None and col_idx < len(rows[total_row_idx]):
            total_votes = _parse_votes(rows[total_row_idx][col_idx]) or 0
        else:
            # 合計行がない場合はデータ行を合算
            total_votes = _sum_data_rows(rows, col_idx, start_row=5)

        candidates.append(
            CandidateRecord(
                name=name,
                party_name=party_name,
                district_name=district_name,
                prefecture=prefecture,
                total_votes=total_votes,
                rank=0,
                is_elected=False,
            )
        )

    # 得票数でソートしてrank付与、最多得票=当選
    candidates.sort(key=lambda c: c.total_votes, reverse=True)
    for i, c in enumerate(candidates):
        c.rank = i + 1
        c.is_elected = i == 0 and c.total_votes > 0

    election_info = None
    if election_date:
        election_info = ElectionInfo(
            election_number=0,
            election_date=election_date,
        )

    return candidates, election_info


def _is_header_cell(value: str) -> bool:
    """セルがヘッダー（候補者名ではない）かどうかを判定する."""
    header_keywords = [
        "候補者",
        "氏名",
        "名前",
        "届出",
        "番号",
        "市区町村",
        "得票数",
        "開票区",
        "選挙区",
        "投票",
    ]
    return any(keyword in value for keyword in header_keywords)
