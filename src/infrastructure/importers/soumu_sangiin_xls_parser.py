"""総務省参議院選挙XLS/XLSXファイルパーサー.

参議院選挙区のXLSファイル構造（1シート=1都道府県/選挙区）:
    Row 1: 選挙日（例: 令和４年７月１０日執行）
    Row 2: タイトル（参議院議員通常選挙（選挙区）候補者別市区町村別得票数一覧）
    Row 3: 選挙区名（例: 北海道）
    Row 4: "候補者名" + 候補者名（各列に1名）+ "得票数計"
    Row 5: "市区町村名＼政党等名" + 政党名
    Row 6+: 市区町村別得票数
    最終データ行: 合計得票数（"合計" or "{都道府県名} 合計"）

衆議院との主な違い:
    - 1シート=1都道府県（1選挙区）
    - 複数当選者がありうる（定数 > 1）
    - 候補者名行の最後に "得票数計" 列がある
"""

import logging

from pathlib import Path

from src.domain.value_objects.election_candidate import CandidateRecord, ElectionInfo
from src.infrastructure.importers._constants import PREFECTURE_NAMES
from src.infrastructure.importers._utils import parse_wareki_date, zen_to_han


logger = logging.getLogger(__name__)

# 合区マッピング（第24回以降: コンポーネント県→合区名）
# 第24回(2016年)から鳥取県・島根県、徳島県・高知県が合区
GOUKU_MAPPING: dict[str, str] = {
    "鳥取県": "鳥取県・島根県",
    "島根県": "鳥取県・島根県",
    "徳島県": "徳島県・高知県",
    "高知県": "徳島県・高知県",
}
# 合区導入開始回次
GOUKU_START_ELECTION = 24

# 参議院選挙区の改選定数（回次→都道府県名→定数）
# 半数改選のため、改選がない選挙区は含まれない
# 出典: 総務省選挙結果 基礎データ
SANGIIN_SEATS: dict[int, dict[str, int]] = {
    27: {
        "北海道": 2,
        "青森県": 1,
        "岩手県": 1,
        "宮城県": 1,
        "秋田県": 1,
        "山形県": 1,
        "福島県": 1,
        "茨城県": 2,
        "栃木県": 1,
        "群馬県": 1,
        "埼玉県": 3,
        "千葉県": 3,
        "東京都": 6,
        "神奈川県": 4,
        "新潟県": 1,
        "富山県": 1,
        "石川県": 1,
        "福井県": 1,
        "山梨県": 1,
        "長野県": 1,
        "岐阜県": 1,
        "静岡県": 2,
        "愛知県": 4,
        "三重県": 1,
        "滋賀県": 1,
        "京都府": 2,
        "大阪府": 4,
        "兵庫県": 3,
        "奈良県": 1,
        "和歌山県": 1,
        "鳥取県・島根県": 1,
        "岡山県": 1,
        "広島県": 2,
        "山口県": 1,
        "徳島県・高知県": 1,
        "香川県": 1,
        "愛媛県": 1,
        "福岡県": 3,
        "佐賀県": 1,
        "長崎県": 1,
        "熊本県": 1,
        "大分県": 1,
        "宮崎県": 1,
        "鹿児島県": 1,
        "沖縄県": 1,
    },
    26: {
        "北海道": 3,
        "青森県": 1,
        "岩手県": 1,
        "宮城県": 1,
        "秋田県": 1,
        "山形県": 1,
        "福島県": 1,
        "茨城県": 2,
        "栃木県": 1,
        "群馬県": 1,
        "埼玉県": 4,
        "千葉県": 3,
        "東京都": 6,
        "神奈川県": 4,
        "新潟県": 1,
        "富山県": 1,
        "石川県": 1,
        "福井県": 1,
        "山梨県": 1,
        "長野県": 1,
        "岐阜県": 1,
        "静岡県": 2,
        "愛知県": 4,
        "三重県": 1,
        "滋賀県": 1,
        "京都府": 2,
        "大阪府": 4,
        "兵庫県": 3,
        "奈良県": 1,
        "和歌山県": 1,
        "鳥取県・島根県": 1,
        "岡山県": 1,
        "広島県": 2,
        "山口県": 1,
        "徳島県・高知県": 1,
        "香川県": 1,
        "愛媛県": 1,
        "福岡県": 3,
        "佐賀県": 1,
        "長崎県": 1,
        "熊本県": 1,
        "大分県": 1,
        "宮崎県": 1,
        "鹿児島県": 1,
        "沖縄県": 1,
    },
    25: {
        "北海道": 3,
        "青森県": 1,
        "岩手県": 1,
        "宮城県": 1,
        "秋田県": 1,
        "山形県": 1,
        "福島県": 1,
        "茨城県": 2,
        "栃木県": 1,
        "群馬県": 1,
        "埼玉県": 3,
        "千葉県": 3,
        "東京都": 6,
        "神奈川県": 4,
        "新潟県": 1,
        "富山県": 1,
        "石川県": 1,
        "福井県": 1,
        "山梨県": 1,
        "長野県": 1,
        "岐阜県": 1,
        "静岡県": 2,
        "愛知県": 3,
        "三重県": 1,
        "滋賀県": 1,
        "京都府": 2,
        "大阪府": 4,
        "兵庫県": 3,
        "奈良県": 1,
        "和歌山県": 1,
        "鳥取県・島根県": 1,
        "岡山県": 1,
        "広島県": 2,
        "山口県": 1,
        "徳島県・高知県": 1,
        "香川県": 1,
        "愛媛県": 1,
        "福岡県": 3,
        "佐賀県": 1,
        "長崎県": 1,
        "熊本県": 1,
        "大分県": 1,
        "宮崎県": 1,
        "鹿児島県": 1,
        "沖縄県": 1,
    },
    24: {
        "北海道": 3,
        "青森県": 1,
        "岩手県": 1,
        "宮城県": 1,
        "秋田県": 1,
        "山形県": 1,
        "福島県": 1,
        "茨城県": 2,
        "栃木県": 1,
        "群馬県": 1,
        "埼玉県": 3,
        "千葉県": 3,
        "東京都": 6,
        "神奈川県": 4,
        "新潟県": 1,
        "富山県": 1,
        "石川県": 1,
        "福井県": 1,
        "山梨県": 1,
        "長野県": 1,
        "岐阜県": 1,
        "静岡県": 2,
        "愛知県": 4,
        "三重県": 1,
        "滋賀県": 1,
        "京都府": 2,
        "大阪府": 4,
        "兵庫県": 3,
        "奈良県": 1,
        "和歌山県": 1,
        "鳥取県・島根県": 1,
        "岡山県": 1,
        "広島県": 2,
        "山口県": 1,
        "徳島県・高知県": 1,
        "香川県": 1,
        "愛媛県": 1,
        "福岡県": 3,
        "佐賀県": 1,
        "長崎県": 1,
        "熊本県": 1,
        "大分県": 1,
        "宮崎県": 1,
        "鹿児島県": 1,
        "沖縄県": 1,
    },
    23: {
        "北海道": 2,
        "青森県": 1,
        "岩手県": 1,
        "宮城県": 2,
        "秋田県": 1,
        "山形県": 1,
        "福島県": 1,
        "茨城県": 2,
        "栃木県": 1,
        "群馬県": 1,
        "埼玉県": 3,
        "千葉県": 3,
        "東京都": 5,
        "神奈川県": 4,
        "新潟県": 2,
        "富山県": 1,
        "石川県": 1,
        "福井県": 1,
        "山梨県": 1,
        "長野県": 2,
        "岐阜県": 2,
        "静岡県": 2,
        "愛知県": 3,
        "三重県": 1,
        "滋賀県": 1,
        "京都府": 2,
        "大阪府": 4,
        "兵庫県": 2,
        "奈良県": 1,
        "和歌山県": 1,
        "鳥取県": 1,
        "島根県": 1,
        "岡山県": 1,
        "広島県": 2,
        "山口県": 1,
        "徳島県": 1,
        "香川県": 1,
        "愛媛県": 1,
        "高知県": 1,
        "福岡県": 2,
        "佐賀県": 1,
        "長崎県": 1,
        "熊本県": 1,
        "大分県": 1,
        "宮崎県": 1,
        "鹿児島県": 1,
        "沖縄県": 1,
    },
    22: {
        "北海道": 2,
        "青森県": 1,
        "岩手県": 1,
        "宮城県": 2,
        "秋田県": 1,
        "山形県": 1,
        "福島県": 1,
        "茨城県": 2,
        "栃木県": 1,
        "群馬県": 1,
        "埼玉県": 3,
        "千葉県": 3,
        "東京都": 5,
        "神奈川県": 4,
        "新潟県": 2,
        "富山県": 1,
        "石川県": 1,
        "福井県": 1,
        "山梨県": 1,
        "長野県": 2,
        "岐阜県": 2,
        "静岡県": 2,
        "愛知県": 3,
        "三重県": 1,
        "滋賀県": 1,
        "京都府": 2,
        "大阪府": 3,
        "兵庫県": 2,
        "奈良県": 1,
        "和歌山県": 1,
        "鳥取県": 1,
        "島根県": 1,
        "岡山県": 1,
        "広島県": 2,
        "山口県": 1,
        "徳島県": 1,
        "香川県": 1,
        "愛媛県": 1,
        "高知県": 1,
        "福岡県": 2,
        "佐賀県": 1,
        "長崎県": 1,
        "熊本県": 1,
        "大分県": 1,
        "宮崎県": 1,
        "鹿児島県": 1,
        "沖縄県": 1,
    },
    21: {
        "北海道": 2,
        "青森県": 1,
        "岩手県": 1,
        "宮城県": 2,
        "秋田県": 1,
        "山形県": 1,
        "福島県": 1,
        "茨城県": 2,
        "栃木県": 1,
        "群馬県": 1,
        "埼玉県": 3,
        "千葉県": 3,
        "東京都": 5,
        "神奈川県": 3,
        "新潟県": 2,
        "富山県": 1,
        "石川県": 1,
        "福井県": 1,
        "山梨県": 1,
        "長野県": 2,
        "岐阜県": 2,
        "静岡県": 2,
        "愛知県": 3,
        "三重県": 1,
        "滋賀県": 1,
        "京都府": 2,
        "大阪府": 3,
        "兵庫県": 2,
        "奈良県": 1,
        "和歌山県": 1,
        "鳥取県": 1,
        "島根県": 1,
        "岡山県": 1,
        "広島県": 2,
        "山口県": 1,
        "徳島県": 1,
        "香川県": 1,
        "愛媛県": 1,
        "高知県": 1,
        "福岡県": 2,
        "佐賀県": 1,
        "長崎県": 1,
        "熊本県": 1,
        "大分県": 1,
        "宮崎県": 1,
        "鹿児島県": 1,
        "沖縄県": 1,
    },
}


def _clean_cell_value(value: object) -> str | None:
    """セルの値を文字列にクリーンアップする."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    return s


def _parse_votes(value: object) -> int | None:
    """得票数セルの値をintに変換する."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = zen_to_han(str(value).strip().replace(",", "").replace("，", ""))
    s = s.replace(".", "").replace(" ", "").replace("\u3000", "")
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _extract_prefecture(district_name: str) -> str:
    """選挙区名から都道府県名を抽出する.

    合区対応: "鳥取県・島根県" のような名前もそのまま返す。
    """
    # 完全一致チェック
    for pref in PREFECTURE_NAMES:
        if district_name == pref:
            return pref
    # 合区チェック（"鳥取県・島根県" 等）
    if "・" in district_name:
        return district_name
    # 部分一致チェック
    for pref in PREFECTURE_NAMES:
        if district_name.startswith(pref):
            return pref
    return district_name


def _sum_data_rows(rows: list[tuple[object, ...]], col_idx: int, start_row: int) -> int:
    """データ行の指定列を合算する（合計行がない場合のフォールバック）."""
    total = 0
    for row_idx in range(start_row, len(rows)):
        row = rows[row_idx]
        if col_idx >= len(row):
            continue
        votes = _parse_votes(row[col_idx])
        if votes is not None:
            total += votes
    return total


def _find_total_row_index(rows: list[tuple[object, ...]], start_row: int) -> int | None:
    """合計行のインデックスを見つける."""
    for i in range(len(rows) - 1, start_row, -1):
        row = rows[i]
        if not row:
            continue
        first_cell = _clean_cell_value(row[0])
        if first_cell and ("合計" in first_cell or first_cell in ("計", "合\u3000計")):
            return i
        if len(row) > 1:
            second_cell = _clean_cell_value(row[1])
            if second_cell and (
                "合計" in second_cell or second_cell in ("計", "合\u3000計")
            ):
                return i
    return None


def _is_skip_column(name: str) -> bool:
    """候補者名ではないスキップすべき列かどうかを判定する."""
    skip_keywords = [
        "候補者",
        "得票数計",
        "合計",
        "計",
        "氏名",
        "市区町村",
        "開票区",
        "投票",
    ]
    return any(keyword in name for keyword in skip_keywords)


def _get_seats_for_district(election_number: int, district_name: str) -> int | None:
    """選挙回次と選挙区名から定数を取得する."""
    seats_map = SANGIIN_SEATS.get(election_number)
    if seats_map is None:
        return None
    seats = seats_map.get(district_name)
    if seats is not None:
        return seats
    # 部分一致で再検索（合区対応）
    for key, val in seats_map.items():
        if district_name in key or key in district_name:
            return val
    return None


def parse_sangiin_xls_file(
    file_path: Path,
    election_number: int | None = None,
) -> tuple[ElectionInfo | None, list[CandidateRecord]]:
    """参議院XLS/XLSXファイルをパースして候補者データを抽出する.

    Args:
        file_path: XLS/XLSXファイルのパス
        election_number: 選挙回次（定数判定に使用、省略時は全員落選扱い）

    Returns:
        (選挙情報, 候補者レコードのリスト)
    """
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _parse_sangiin_xlsx(file_path, election_number)
    elif suffix == ".xls":
        return _parse_sangiin_xls(file_path, election_number)
    else:
        logger.error("未対応のファイル形式: %s", suffix)
        return None, []


def _parse_sangiin_xlsx(
    file_path: Path,
    election_number: int | None = None,
) -> tuple[ElectionInfo | None, list[CandidateRecord]]:
    """openpyxlを使用して.xlsxファイルをパースする."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    # 選挙区名→候補者リストのマップ（同一選挙区名の重複排除用）
    # 合区正規化により複数シートが同名になる場合、最多候補者のシートを採用
    district_candidates: dict[str, list[CandidateRecord]] = {}
    election_info: ElectionInfo | None = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[tuple[object, ...]] = [
            tuple(cell for cell in row) for row in ws.iter_rows(values_only=True)
        ]
        if len(rows) < 5:
            logger.debug("シート '%s' のデータ行が不足: %d行", sheet_name, len(rows))
            continue

        candidates, sheet_election_info = _parse_sangiin_rows(
            rows, election_number, sheet_name=sheet_name
        )
        if sheet_election_info and election_info is None:
            election_info = sheet_election_info

        if not candidates:
            continue
        district = candidates[0].district_name
        existing = district_candidates.get(district)
        if existing is None or len(candidates) > len(existing):
            district_candidates[district] = candidates

    wb.close()
    all_candidates: list[CandidateRecord] = []
    for candidates in district_candidates.values():
        all_candidates.extend(candidates)
    return election_info, all_candidates


def _parse_sangiin_xls(
    file_path: Path,
    election_number: int | None = None,
) -> tuple[ElectionInfo | None, list[CandidateRecord]]:
    """xlrdを使用して.xlsファイルをパースする."""
    import xlrd

    wb = xlrd.open_workbook(str(file_path))
    district_candidates: dict[str, list[CandidateRecord]] = {}
    election_info: ElectionInfo | None = None

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        if ws.nrows < 5:
            logger.debug("シート '%s' のデータ行が不足: %d行", ws.name, ws.nrows)
            continue

        rows: list[tuple[object, ...]] = []
        for row_idx in range(ws.nrows):
            row: tuple[object, ...] = tuple(
                ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)
            )
            rows.append(row)

        candidates, sheet_election_info = _parse_sangiin_rows(
            rows, election_number, sheet_name=ws.name
        )
        if sheet_election_info and election_info is None:
            election_info = sheet_election_info

        if not candidates:
            continue
        district = candidates[0].district_name
        existing = district_candidates.get(district)
        if existing is None or len(candidates) > len(existing):
            district_candidates[district] = candidates

    all_candidates: list[CandidateRecord] = []
    for candidates in district_candidates.values():
        all_candidates.extend(candidates)
    return election_info, all_candidates


def _parse_sangiin_rows(
    rows: list[tuple[object, ...]],
    election_number: int | None = None,
    sheet_name: str | None = None,
) -> tuple[list[CandidateRecord], ElectionInfo | None]:
    """行データから参議院候補者レコードを抽出する.

    Args:
        rows: シートの全行データ
        election_number: 選挙回次（定数判定用）
        sheet_name: シート名（選挙区名のフォールバック用）

    Returns:
        (候補者リスト, 選挙情報)
    """
    if len(rows) < 5:
        return [], None

    # Row 1: 選挙日
    election_date_str = _clean_cell_value(rows[0][0]) if rows[0] else None
    election_date = parse_wareki_date(election_date_str or "")

    # Row 3: 選挙区名（0-indexed: rows[2]）
    # data_only=Trueで読んでも数式の結果がキャッシュされていない場合がある
    district_name = ""
    for cell in rows[2]:
        val = _clean_cell_value(cell)
        if val and val != "[単位：票]" and not val.startswith("="):
            district_name = zen_to_han(val)
            break

    # シート名をフォールバックとして使用（XLSファイルのシート名は都道府県名）
    if not district_name and sheet_name:
        district_name = zen_to_han(sheet_name)

    if not district_name:
        logger.debug("選挙区名が取得できません")
        return [], None

    # 合区正規化（第24回以降: 鳥取県→鳥取県・島根県 等）
    if election_number is not None and election_number >= GOUKU_START_ELECTION:
        normalized = GOUKU_MAPPING.get(district_name)
        if normalized:
            district_name = normalized

    # タイトル確認（参議院選挙区データかチェック）
    title_row = _clean_cell_value(rows[1][1]) if len(rows[1]) > 1 else None
    if title_row and "比例代表" in title_row:
        logger.debug("比例代表データをスキップ: %s", district_name)
        return [], None

    prefecture = _extract_prefecture(district_name)

    # Row 4-5: 候補者名と政党名（0-indexed: rows[3], rows[4]）
    name_row = rows[3]
    party_row = rows[4]

    # 候補者データの列を特定（名前が入っている列、スキップ列を除外）
    candidate_columns: list[int] = []
    for col_idx in range(len(name_row)):
        name = _clean_cell_value(name_row[col_idx])
        if name and not _is_skip_column(name):
            candidate_columns.append(col_idx)

    if not candidate_columns:
        logger.debug("候補者が見つかりません: %s", district_name)
        return [], None

    # 合計行を見つける
    total_row_idx = _find_total_row_index(rows, 5)

    # 候補者データを抽出
    candidates: list[CandidateRecord] = []
    for col_idx in candidate_columns:
        name = _clean_cell_value(name_row[col_idx])
        if not name:
            continue

        party_name = ""
        if col_idx < len(party_row):
            party_name = _clean_cell_value(party_row[col_idx]) or ""

        # 合計得票数
        total_votes = 0
        if total_row_idx is not None and col_idx < len(rows[total_row_idx]):
            total_votes = _parse_votes(rows[total_row_idx][col_idx]) or 0
        else:
            total_votes = _sum_data_rows(rows, col_idx, start_row=5)

        candidates.append(
            CandidateRecord(
                name=name,
                party_name=party_name,
                district_name=district_name,
                prefecture=prefecture,
                total_votes=total_votes,
                rank=0,
                is_elected=False,
            )
        )

    # 得票数でソートしてrank付与
    candidates.sort(key=lambda c: c.total_votes, reverse=True)
    seats = (
        _get_seats_for_district(election_number, district_name)
        if election_number is not None
        else None
    )

    for i, c in enumerate(candidates):
        c.rank = i + 1
        if seats is not None:
            c.is_elected = i < seats and c.total_votes > 0
        else:
            # 定数不明の場合は最多得票1名のみ当選とする
            c.is_elected = i == 0 and c.total_votes > 0

    election_info = None
    if election_date:
        election_info = ElectionInfo(
            election_number=election_number or 0,
            election_date=election_date,
        )

    return candidates, election_info
