"""総務省比例代表XLSファイルパーサー.

第48回（2017年）のXLSファイル（1423行×28列）をパースする。
シート内にブロック単位でセクションが分割されている構造に対応。

構造概要:
    - 1シートに全11ブロック分のデータが格納
    - ブロック名（「北海道ブロック」等）がセクション見出し
    - 各ブロック内で政党ごとに候補者一覧
    - 候補者行: 名簿順位、候補者名、小選挙区結果、惜敗率 等
"""

import logging
import re

from datetime import date
from pathlib import Path

from src.domain.value_objects.proportional_candidate import (
    ProportionalCandidateRecord,
    ProportionalElectionInfo,
)
from src.infrastructure.importers._constants import PROPORTIONAL_BLOCKS


# 和暦→西暦変換
_WAREKI_MAP = {
    "令和": 2018,
    "平成": 1988,
    "昭和": 1925,
}


logger = logging.getLogger(__name__)

# ブロック名の検出パターン
_BLOCK_PATTERN = re.compile(
    r"(北海道|東北|北関東|南関東|東京|北陸信越|東海|近畿|中国|四国|九州)\s*ブロック"
)

# 当選人数を示すパターン
_WINNERS_PATTERN = re.compile(r"当選人?\s*(\d+)")


def _zen_to_han(text: str) -> str:
    """全角数字・記号を半角に変換する."""
    zen = "０１２３４５６７８９．"
    han = "0123456789."
    table = str.maketrans(zen, han)
    return text.translate(table)


def _parse_wareki_date(text: str) -> date | None:
    """和暦の日付文字列を西暦dateに変換する."""
    if not text:
        return None
    text = _zen_to_han(str(text))
    pattern = r"(令和|平成|昭和)(\d+)年(\d+)月(\d+)日"
    match = re.search(pattern, text)
    if not match:
        return None
    era, year_str, month_str, day_str = match.groups()
    base_year = _WAREKI_MAP.get(era)
    if base_year is None:
        return None
    year = base_year + int(year_str)
    return date(year, int(month_str), int(day_str))


def _clean_cell(value: object) -> str:
    """セル値を文字列にクリーンアップする."""
    if value is None:
        return ""
    s = str(value).strip()
    if s == "":
        return ""
    return _zen_to_han(s)


def _parse_float(value: object) -> float | None:
    """セル値をfloatに変換する."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value != 0 else None
    s = _zen_to_han(str(value).strip().replace(",", "").replace("，", ""))
    s = s.replace("%", "").replace("％", "")
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_int(value: object) -> int | None:
    """セル値をintに変換する."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value) if value != 0 else None
    s = _zen_to_han(str(value).strip().replace(",", "").replace("，", ""))
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _detect_block_name(row: tuple[object, ...]) -> str | None:
    """行からブロック名を検出する."""
    for cell in row:
        s = _clean_cell(cell)
        if not s:
            continue
        match = _BLOCK_PATTERN.search(s)
        if match:
            return match.group(1)
        # 単純なブロック名一致もチェック
        for block in PROPORTIONAL_BLOCKS:
            if s == block or s == f"{block}ブロック":
                return block
    return None


def parse_proportional_xls(
    file_path: Path,
) -> tuple[ProportionalElectionInfo | None, list[ProportionalCandidateRecord]]:
    """比例代表XLSファイルをパースする.

    Args:
        file_path: XLSファイルのパス

    Returns:
        (選挙情報, 比例代表候補者レコードのリスト)
    """
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx_proportional(file_path)
    elif suffix == ".xls":
        return _parse_xls_proportional(file_path)
    else:
        logger.error("未対応のファイル形式: %s", suffix)
        return None, []


def _parse_xlsx_proportional(
    file_path: Path,
) -> tuple[ProportionalElectionInfo | None, list[ProportionalCandidateRecord]]:
    """openpyxlを使用して.xlsxファイルをパースする."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    all_candidates: list[ProportionalCandidateRecord] = []
    election_info: ProportionalElectionInfo | None = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[tuple[object, ...]] = [
            tuple(cell for cell in row) for row in ws.iter_rows(values_only=True)
        ]
        info, candidates = _parse_proportional_rows(rows)
        if info and election_info is None:
            election_info = info
        all_candidates.extend(candidates)

    wb.close()
    return election_info, all_candidates


def _parse_xls_proportional(
    file_path: Path,
) -> tuple[ProportionalElectionInfo | None, list[ProportionalCandidateRecord]]:
    """xlrdを使用して.xlsファイルをパースする."""
    import xlrd

    wb = xlrd.open_workbook(str(file_path))
    all_candidates: list[ProportionalCandidateRecord] = []
    election_info: ProportionalElectionInfo | None = None

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        if ws.nrows < 3:
            continue

        rows: list[tuple[object, ...]] = []
        for row_idx in range(ws.nrows):
            row: tuple[object, ...] = tuple(
                ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)
            )
            rows.append(row)

        info, candidates = _parse_proportional_rows(rows)
        if info and election_info is None:
            election_info = info
        all_candidates.extend(candidates)

    return election_info, all_candidates


def _find_column_layout(
    rows: list[tuple[object, ...]], start_idx: int
) -> dict[str, int] | None:
    """ヘッダー行からカラムレイアウトを検出する.

    Returns:
        カラム名→列番号のdict、またはNone
    """
    for i in range(start_idx, min(start_idx + 10, len(rows))):
        row = rows[i]
        layout: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            s = _clean_cell(cell)
            if not s:
                continue
            if "順位" in s or "名簿" in s:
                layout["list_order"] = col_idx
            elif "氏名" in s or "候補者" in s or "名前" in s:
                layout["name"] = col_idx
            elif "小選挙区" in s or "選挙区" in s:
                layout["smd_result"] = col_idx
            elif "惜敗" in s or "率" in s:
                layout["loss_ratio"] = col_idx
            elif "当" in s and "落" in s:
                layout["smd_result"] = col_idx

        if "name" in layout:
            return layout
    return None


def _parse_proportional_rows(
    rows: list[tuple[object, ...]],
) -> tuple[ProportionalElectionInfo | None, list[ProportionalCandidateRecord]]:
    """行データから比例代表候補者レコードを抽出する.

    比例代表XLSは以下の構造:
    - ブロック見出し行（「北海道ブロック」等）
    - 政党名見出し行
    - カラムヘッダー行（順位、候補者名、小選挙区結果等）
    - 候補者データ行
    """
    if len(rows) < 3:
        return None, []

    all_candidates: list[ProportionalCandidateRecord] = []
    election_info: ProportionalElectionInfo | None = None

    # 選挙日を最初の数行から抽出
    for row in rows[:5]:
        for cell in row:
            s = _clean_cell(cell)
            if s:
                d = _parse_wareki_date(s)
                if d:
                    election_info = ProportionalElectionInfo(
                        election_number=0,
                        election_date=d,
                    )
                    break
        if election_info:
            break

    current_block: str | None = None
    current_party: str | None = None
    current_layout: dict[str, int] | None = None
    current_winners_count: int = 0
    party_candidate_count: int = 0

    i = 0
    while i < len(rows):
        row = rows[i]

        # ブロック検出
        block = _detect_block_name(row)
        if block:
            current_block = block
            current_party = None
            current_layout = None
            logger.debug("ブロック検出: %s (行 %d)", block, i + 1)
            i += 1
            continue

        if current_block is None:
            i += 1
            continue

        # 行のテキスト内容を分析
        row_text = " ".join(_clean_cell(c) for c in row).strip()

        # 空行はスキップ
        if not row_text:
            i += 1
            continue

        # 政党名行の検出:
        # 政党名は通常、非数値テキストで始まり、候補者データ行ではない
        # 最初の非空セルが政党名候補
        first_non_empty = ""
        for cell in row:
            s = _clean_cell(cell)
            if s:
                first_non_empty = s
                break

        # 当選人数を含む行は政党ヘッダーの一部
        winners_match = _WINNERS_PATTERN.search(row_text)

        # レイアウト検出の試行
        if current_party and current_layout is None:
            layout = _find_column_layout(rows, i)
            if layout:
                current_layout = layout
                i += 1
                continue

        # 候補者データ行の判定
        if current_party and current_layout:
            name_col = current_layout.get("name", 1)
            if name_col < len(row):
                name = _clean_cell(row[name_col])
                if name and not any(
                    kw in name
                    for kw in [
                        "候補者",
                        "氏名",
                        "名簿",
                        "順位",
                        "合計",
                        "計",
                        "政党",
                        "ブロック",
                    ]
                ):
                    # 候補者データ行
                    list_order_col = current_layout.get("list_order", 0)
                    smd_col = current_layout.get("smd_result", -1)
                    loss_col = current_layout.get("loss_ratio", -1)

                    list_order = (
                        _parse_int(row[list_order_col])
                        if list_order_col < len(row)
                        else None
                    )
                    smd_result = ""
                    if 0 <= smd_col < len(row):
                        smd_result = _clean_cell(row[smd_col])

                    loss_ratio = None
                    if 0 <= loss_col < len(row):
                        loss_ratio = _parse_float(row[loss_col])

                    party_candidate_count += 1
                    is_elected = party_candidate_count <= current_winners_count

                    # 比例復活 vs 比例当選の判定
                    # smd_result: "当"=小選挙区当選, "落"=小選挙区落選, ""=比例単独
                    candidate = ProportionalCandidateRecord(
                        name=name,
                        party_name=current_party,
                        block_name=current_block,
                        list_order=list_order or party_candidate_count,
                        smd_result=smd_result,
                        loss_ratio=loss_ratio,
                        is_elected=is_elected,
                    )
                    all_candidates.append(candidate)
                    i += 1
                    continue

        # 政党名行の検出（非数値、非ヘッダー、2文字以上）
        if (
            first_non_empty
            and len(first_non_empty) >= 2
            and not first_non_empty[0].isdigit()
            and not any(
                kw in first_non_empty
                for kw in ["順位", "氏名", "候補者", "合計", "計", "当選"]
            )
        ):
            # 新しい政党セクション
            potential_party = first_non_empty
            # 当選人数を同じ行から抽出
            winners = 0
            if winners_match:
                winners = int(winners_match.group(1))

            current_party = potential_party
            current_winners_count = winners
            current_layout = None
            party_candidate_count = 0
            logger.debug(
                "政党検出: %s (当選%d名, ブロック=%s, 行 %d)",
                current_party,
                current_winners_count,
                current_block,
                i + 1,
            )

            # 次の行でレイアウトを検出
            if i + 1 < len(rows):
                layout = _find_column_layout(rows, i + 1)
                if layout:
                    current_layout = layout
                    i += 2  # 政党行 + ヘッダー行をスキップ
                    continue

        i += 1

    logger.info("比例代表XLSパース完了: %d候補者", len(all_candidates))
    return election_info, all_candidates


def get_elected_candidates(
    candidates: list[ProportionalCandidateRecord],
) -> list[ProportionalCandidateRecord]:
    """当選者のみを抽出する."""
    return [c for c in candidates if c.is_elected]
