"""総務省選挙XLSパーサーのテスト."""

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from src.infrastructure.importers.soumu_xls_parser import (
    _extract_prefecture,
    _parse_votes,
    _parse_wareki_date,
    _zen_to_han,
    parse_xls_file,
)


class TestZenToHan:
    """全角→半角変換のテスト."""

    def test_full_width_numbers(self) -> None:
        assert _zen_to_han("１２３") == "123"

    def test_mixed(self) -> None:
        assert _zen_to_han("第１区") == "第1区"

    def test_no_conversion_needed(self) -> None:
        assert _zen_to_han("abc123") == "abc123"


class TestParseWarekiDate:
    """和暦日付パースのテスト."""

    def test_reiwa(self) -> None:
        result = _parse_wareki_date("令和６年１０月２７日執行")
        assert result == date(2024, 10, 27)

    def test_reiwa_han(self) -> None:
        result = _parse_wareki_date("令和3年10月31日執行")
        assert result == date(2021, 10, 31)

    def test_heisei(self) -> None:
        result = _parse_wareki_date("平成29年10月22日執行")
        assert result == date(2017, 10, 22)

    def test_invalid(self) -> None:
        assert _parse_wareki_date("invalid") is None

    def test_empty(self) -> None:
        assert _parse_wareki_date("") is None


class TestExtractPrefecture:
    """都道府県抽出のテスト."""

    def test_hokkaido(self) -> None:
        assert _extract_prefecture("北海道第１区") == "北海道"

    def test_tokyo(self) -> None:
        assert _extract_prefecture("東京都第5区") == "東京都"

    def test_osaka(self) -> None:
        assert _extract_prefecture("大阪府第3区") == "大阪府"

    def test_unknown(self) -> None:
        assert _extract_prefecture("不明区") == ""


class TestParseVotes:
    """得票数パースのテスト."""

    def test_integer(self) -> None:
        assert _parse_votes(12345) == 12345

    def test_float(self) -> None:
        assert _parse_votes(12345.0) == 12345

    def test_string_with_comma(self) -> None:
        assert _parse_votes("12,345") == 12345

    def test_full_width(self) -> None:
        assert _parse_votes("１２３４５") == 12345

    def test_none(self) -> None:
        assert _parse_votes(None) is None

    def test_empty(self) -> None:
        assert _parse_votes("") is None


class TestParseXlsxFile:
    """XLSXファイルパースのテスト."""

    @pytest.fixture()
    def sample_xlsx(self, tmp_path: Path) -> Path:
        """テスト用のXLSXファイルを生成する."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"

        # Row 1: 選挙日
        ws.append(["令和６年１０月２７日執行"])
        # Row 2: タイトル
        ws.append(["衆議院議員総選挙（小選挙区）候補者別得票数"])
        # Row 3: 選挙区名
        ws.append(["北海道第１区", None, None])
        # Row 4: 候補者名
        ws.append([None, "候補A太郎", "候補B次郎", "候補C三郎"])
        # Row 5: 政党名
        ws.append([None, "自由民主党", "立憲民主党", "日本共産党"])
        # Row 6: 市区町村別得票数
        ws.append(["札幌市中央区", 30000, 25000, 10000])
        ws.append(["札幌市北区", 20000, 15000, 8000])
        # Row 8: 合計
        ws.append(["合計", 50000, 40000, 18000])

        file_path = tmp_path / "test_election.xlsx"
        wb.save(file_path)
        wb.close()
        return file_path

    def test_parse_basic_xlsx(self, sample_xlsx: Path) -> None:
        """基本的なXLSXファイルのパースを確認する."""
        election_info, candidates = parse_xls_file(sample_xlsx)

        assert election_info is not None
        assert election_info.election_date == date(2024, 10, 27)

        assert len(candidates) == 3

        # 得票数降順でソートされている
        assert candidates[0].name == "候補A太郎"
        assert candidates[0].total_votes == 50000
        assert candidates[0].rank == 1
        assert candidates[0].is_elected is True
        assert candidates[0].party_name == "自由民主党"
        assert candidates[0].prefecture == "北海道"
        assert candidates[0].district_name == "北海道第1区"

        assert candidates[1].name == "候補B次郎"
        assert candidates[1].total_votes == 40000
        assert candidates[1].rank == 2
        assert candidates[1].is_elected is False

        assert candidates[2].name == "候補C三郎"
        assert candidates[2].total_votes == 18000
        assert candidates[2].rank == 3
        assert candidates[2].is_elected is False

    @pytest.fixture()
    def multi_sheet_xlsx(self, tmp_path: Path) -> Path:
        """複数シート（複数選挙区）のXLSXファイルを生成する."""
        wb = openpyxl.Workbook()

        # 第1シート: 北海道第1区
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "北海道第1区"
        ws1.append(["令和６年１０月２７日執行"])
        ws1.append(["タイトル"])
        ws1.append(["北海道第１区"])
        ws1.append([None, "候補A", "候補B"])
        ws1.append([None, "党A", "党B"])
        ws1.append(["地域1", 30000, 20000])
        ws1.append(["合計", 30000, 20000])

        # 第2シート: 北海道第2区
        ws2 = wb.create_sheet("北海道第2区")
        ws2.append(["令和６年１０月２７日執行"])
        ws2.append(["タイトル"])
        ws2.append(["北海道第２区"])
        ws2.append([None, "候補C", "候補D"])
        ws2.append([None, "党C", "党D"])
        ws2.append(["地域2", 25000, 35000])
        ws2.append(["合計", 25000, 35000])

        file_path = tmp_path / "multi_sheet.xlsx"
        wb.save(file_path)
        wb.close()
        return file_path

    def test_parse_multi_sheet(self, multi_sheet_xlsx: Path) -> None:
        """複数シートのパースを確認する."""
        election_info, candidates = parse_xls_file(multi_sheet_xlsx)

        assert election_info is not None
        assert len(candidates) == 4

        # 第1区の候補者
        hokkaido1 = [c for c in candidates if "1区" in c.district_name]
        assert len(hokkaido1) == 2
        assert hokkaido1[0].name == "候補A"
        assert hokkaido1[0].is_elected is True

        # 第2区の候補者
        hokkaido2 = [c for c in candidates if "2区" in c.district_name]
        assert len(hokkaido2) == 2
        assert hokkaido2[0].name == "候補D"
        assert hokkaido2[0].is_elected is True

    @pytest.fixture()
    def empty_xlsx(self, tmp_path: Path) -> Path:
        """空のXLSXファイルを生成する."""
        wb = openpyxl.Workbook()
        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)
        wb.close()
        return file_path

    def test_parse_empty_xlsx(self, empty_xlsx: Path) -> None:
        """空のXLSXファイルのパースでエラーにならないことを確認する."""
        election_info, candidates = parse_xls_file(empty_xlsx)
        assert election_info is None
        assert len(candidates) == 0

    def test_unsupported_format(self, tmp_path: Path) -> None:
        """未対応の拡張子でNoneが返ることを確認する."""
        file_path = tmp_path / "test.csv"
        file_path.write_text("data")
        election_info, candidates = parse_xls_file(file_path)
        assert election_info is None
        assert len(candidates) == 0
